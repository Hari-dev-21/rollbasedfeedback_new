from rest_framework import viewsets, status, permissions, generics, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from django.db.models import Count, Avg, Q, FloatField
from django.db.models.functions import Cast
from django.db.models.fields.json import KeyTransform

from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
from django.contrib.auth.models import User
# from django.contrib.auth.models import AbstractUser
import json
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, Series
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.lib import colors
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics.shapes import String, Group




from .models import (
    FeedbackForm, Question, FeedbackResponse, Answer, 
    FormAnalytics, Notification, CustomUser, QuestionOption, Section
)
from .serializers import (
    FeedbackFormSerializer, FeedbackFormCreateSerializer,QuestionCreateSerializer,
    FeedbackResponseSerializer, FeedbackResponseCreateSerializer,
    FormAnalyticsSerializer, NotificationSerializer,
    QuestionAnalyticsSerializer, FormSummarySerializer, AdminSerializers, QuestionOptionCreateSerializer, RegisterSerializer,SectionCreateSerializer
)

from .permissions import IsSuperUser
from .consumers import send_notification_to_group

from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.authtoken.models import Token
from rest_framework.permissions import AllowAny
from django.contrib.auth import logout as django_logout
from rest_framework.decorators import api_view, permission_classes


class PendingUsersView(generics.ListAPIView):
    queryset = CustomUser.objects.filter(is_approved=False, is_superuser=False).order_by('id')
    serializer_class = RegisterSerializer
    permission_classes = [permissions.IsAdminUser]  # only superuser


# class ApproveUserView(generics.UpdateAPIView):
#     queryset = CustomUser.objects.all()
#     permission_classes = [permissions.IsAdminUser]
#     serializer_class = RegisterSerializer

#     def patch(self, request, *args, **kwargs):
#         user = self.get_object()
#         user.is_approved = True
#         user.save()
#         return Response({'message': 'User approved successfully'}, status=status.HTTP_200_OK)

class ApproveUserView(generics.UpdateAPIView):
    queryset = CustomUser.objects.all()
    permission_classes = [permissions.IsAdminUser]
    serializer_class = RegisterSerializer

    def patch(self, request, *args, **kwargs):
        user = self.get_object()
        action = request.data.get('action', 'approve')  # Default to approve
        
        if action == 'approve':
            user.is_approved = True
            user.save()
            return Response({'message': 'User approved successfully'}, status=status.HTTP_200_OK)
        elif action == 'reject':
            # Option 1: Delete the user
            username = user.username
            user.delete()
            return Response({'message': f'User "{username}" rejected and deleted successfully'}, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Invalid action. Use "approve" or "reject".'}, status=status.HTTP_400_BAD_REQUEST)


class RegisterView(generics.CreateAPIView):
    queryset = CustomUser.objects.all()
    serializer_class = RegisterSerializer
    permission_classes = []  # Allow anyone to register


class AdminViewset(viewsets.ModelViewSet):

    queryset = CustomUser.objects.filter(is_superuser=False)
    serializer_class = AdminSerializers
    permission_classes = [IsSuperUser]
    lookup_field = 'username'
    
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_admins_list(request):
    """Get list of all admins who have created forms"""
    try:
        # Get distinct users who have created forms
        admin_users = CustomUser.objects.filter(
            created_forms__isnull=False
        ).distinct().values('id', 'username', 'email')
        
        return Response(list(admin_users))
    except Exception as e:
        return Response(
            {'error': f'Unable to load admin list: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

class SectionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing form sections.
    Each section belongs to a FeedbackForm.
    """
    queryset = Section.objects.all().select_related('form').prefetch_related('questions')
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = SectionCreateSerializer

    # def get_serializer_class(self):
    #     if self.action in ['create', 'update', 'partial_update']:
    #         return SectionCreateSerializer
    #     return SectionSerializer

    def get_queryset(self):
        user = self.request.user
        return Section.objects.filter(form__created_by=user)

    def perform_create(self, serializer):
        form_id = self.request.data.get('form')
        try:
            form = FeedbackForm.objects.get(id=form_id, created_by=self.request.user)
            serializer.save(form=form)
        except FeedbackForm.DoesNotExist:
            raise serializers.ValidationError({"error": "Invalid form ID or access denied"})

class QuestionOptionViewSet(viewsets.ModelViewSet):
    queryset = QuestionOption.objects.all().select_related('question', 'next_section')
    serializer_class = QuestionOptionCreateSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        return QuestionOption.objects.filter(question__section__form__created_by=user)

    def perform_create(self, serializer):
        serializer.save()


class QuestionViewSet(viewsets.ModelViewSet):
    """ViewSet for managing questions within sections"""
    queryset = Question.objects.all().select_related('section')
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = QuestionCreateSerializer

    # def get_serializer_class(self):
    #     if self.action in ['create', 'update', 'partial_update']:
    #         return QuestionCreateSerializer
    #     return QuestionSerializer

    def get_queryset(self):
        user = self.request.user
        return Question.objects.filter(section__form__created_by=user)

    def perform_create(self, serializer):
        section_id = self.request.data.get('section')
        try:
            section = Section.objects.get(id=section_id, form__created_by=self.request.user)
            serializer.save(section=section)
        except Section.DoesNotExist:
            raise serializers.ValidationError({"error": "Invalid section ID or access denied"})




class FeedbackFormViewSet(viewsets.ModelViewSet):
    """ViewSet for managing feedback forms"""
    queryset = FeedbackForm.objects.all()
    serializer_class = FeedbackFormSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user

        # If superuser, show all forms and allow filtering by admin_id
        # if user.is_superuser:
        #     queryset = FeedbackForm.objects.all()


        #     admin_name = self.request.query_params.get('admin_name')
        #     if admin_name:
        #         queryset = queryset.filter(created_by__username=admin_name)
        #     return queryset

        return FeedbackForm.objects.filter(created_by=user)
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return FeedbackFormCreateSerializer
        return FeedbackFormSerializer

    # def create(self, request, *args, **kwargs):
    #     try:
    #         with transaction.atomic():
    #             return super().create(request, *args, **kwargs)
    #     except Exception as e:
    #         return Response(
    #             {"error": f"Failed to create form: {str(e)}"}, 
    #             status=status.HTTP_400_BAD_REQUEST
    #         )

    def create(self, request, *args, **kwargs):
            print("🎯 CREATE FORM REQUEST RECEIVED")
            print(f"📦 Request data type: {type(request.data)}")
            print(f"📦 Request data: {request.data}")
            print(f"👤 User: {request.user}")
            print(f"🔑 Authenticated: {request.user.is_authenticated}")
            
            # Check if serializer class is correct
            serializer_class = self.get_serializer_class()
            print(f"📝 Using serializer: {serializer_class.__name__}")
            
            serializer = serializer_class(data=request.data, context={'request': request})
            
            print("🔍 Checking serializer validity...")
            is_valid = serializer.is_valid()
            print(f"✅ Serializer valid: {is_valid}")
            
            if not is_valid:
                print("❌ SERIALIZER VALIDATION ERRORS:")
                print(serializer.errors)
                return Response(
                    {
                        "error": "Validation failed",
                        "details": serializer.errors,
                        "received_data": str(request.data)[:500]  # First 500 chars
                    }, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            print("💾 Saving form...")
            try:
                instance = serializer.save()
                print(f"🎉 Form created successfully! ID: {instance.id}")
                print(f"📊 Form title: {instance.title}")
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                print(f"💥 Error during save: {str(e)}")
                import traceback
                traceback.print_exc()
                return Response(
                    {"error": f"Form creation failed: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
    
    
    def perform_create(self, serializer):
        form = serializer.save()
        
        # Create analytics record
        FormAnalytics.objects.create(form=form)
        
        # Create notification record in database
        Notification.objects.create(
            user=self.request.user,
            notification_type='form_created',
            title='Form Created',
            message=f'Form "{form.title}" created successfully',
            data={"form_id": str(form.id)}
        )
        
        # Send notification
        send_notification_to_group(
            f"user_{self.request.user.id}",
            "form_created",
            f"Form '{form.title}' created successfully",
            {"form_id": str(form.id)}
        )
    
    @action(detail=True, methods=['get'])
    def analytics(self, request, pk=None):
        """Get detailed analytics for a specific form"""
        try:
            form = self.get_object()
            analytics, created = FormAnalytics.objects.get_or_create(form=form)
            analytics.update_analytics()
            
            serializer = FormAnalyticsSerializer(analytics)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {'error': f'Unable to load analytics: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    # @action(detail=True, methods=['get'])
    # def question_analytics(self, request, pk=None):
        """Get analytics for individual questions"""
        try:
            form = self.get_object()
            questions = form.questions.all()
            
            question_analytics = []
            for question in questions:
                answers = Answer.objects.filter(question=question)
                response_count = answers.count()
                
                analytics_data = {
                    'question_id': question.id,
                    'question_text': question.text,
                    'question_type': question.question_type,
                    'response_count': response_count,
                    'average_rating': None,
                    'answer_distribution': {},
                    'options': question.options or []
                }
                
                if question.question_type in ['rating', 'rating_10']:
                    avg_rating = answers.filter(answer_text__regex=r'^\d+$').aggregate(
                        avg=Avg('answer_text')
                    )['avg']
                    analytics_data['average_rating'] = float(avg_rating) if avg_rating else None
                    
                    # Calculate distribution for rating questions
                    max_rating = 10 if question.question_type == 'rating_10' else 5
                    distribution = {}
                    for i in range(1, max_rating + 1):
                        count = answers.filter(answer_text=str(i)).count()
                        distribution[str(i)] = count
                    analytics_data['answer_distribution'] = distribution
                
                elif question.question_type in ['radio', 'checkbox', 'yes_no']:
                    if question.question_type == 'checkbox':
                        # For checkbox questions, handle comma-separated values
                        distribution = {}
                        for answer in answers:
                            # Split comma-separated values and count each option
                            options = [opt.strip() for opt in answer.answer_text.split(',') if opt.strip()]
                            for option in options:
                                distribution[option] = distribution.get(option, 0) + 1
                        analytics_data['answer_distribution'] = distribution
                    else:
                        # For radio and yes_no questions, use the original logic
                        distribution = answers.values('answer_text').annotate(
                            count=Count('answer_text')
                        ).order_by('-count')
                        analytics_data['answer_distribution'] = {
                            item['answer_text']: item['count'] for item in distribution
                        }
                
                question_analytics.append(analytics_data)
            
            serializer = QuestionAnalyticsSerializer(question_analytics, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {'error': f'Unable to load question analytics: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def share_link(self, request, pk=None):
        """Get shareable link for the form"""
        form = self.get_object()
        return Response({
            'shareable_link': form.shareable_link,
            'form_id': str(form.id)
        })
    
    @action(detail=True, methods=['get'])
    def responses(self, request, pk=None):
        """Get all responses for a specific form"""
        try:
            form = self.get_object()
            responses = FeedbackResponse.objects.filter(form=form).order_by('-submitted_at')
            serializer = FeedbackResponseSerializer(responses, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {'error': f'Unable to load responses: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        """Export form responses to Excel"""
        try:
            form = self.get_object()
            responses = FeedbackResponse.objects.filter(form=form).order_by('-submitted_at')

            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{form.title} Responses"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Get all questions for this form
            questions = form.questions.all().order_by('order')

            # Create headers
            headers = ['Response ID', 'Submitted At', 'IP Address']
            for question in questions:
                headers.append(f"{question.text} ({question.question_type})")

            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Write data
            for row_num, response in enumerate(responses, 2):
                ws.cell(row=row_num, column=1, value=str(response.id))
                ws.cell(row=row_num, column=2, value=response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_num, column=3, value=response.ip_address or 'N/A')

                # Create a mapping of question_id to answer for this response
                answer_map = {}
                for answer in response.answers.all():
                    answer_map[answer.question.id] = answer.answer_text

                # Fill in answers for each question
                for col_num, question in enumerate(questions, 4):
                    answer_text = answer_map.get(question.id, 'No Answer')
                    ws.cell(row=row_num, column=col_num, value=answer_text)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Create HTTP response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{form.title}_responses.xlsx"'

            # Save workbook to response
            wb.save(response)
            return response

        except Exception as e:
            return Response(
                {'error': f'Unable to export responses: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def export_csv(self, request, pk=None):
        """Export form responses to CSV"""
        try:
            form = self.get_object()
            responses = FeedbackResponse.objects.filter(form=form).order_by('-submitted_at')

            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{form.title}_responses.csv"'

            # Create CSV writer
            writer = csv.writer(response)

            # Get all questions for this form
            questions = form.questions.all().order_by('order')

            # Write header row
            headers = ['Response ID', 'Submitted At', 'IP Address']
            for question in questions:
                headers.append(f"{question.text} ({question.question_type})")
            writer.writerow(headers)

            # Write data rows
            for feedback_response in responses:
                row = [
                    str(feedback_response.id),
                    feedback_response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
                    feedback_response.ip_address or 'N/A'
                ]

                # Add answers for each question
                for question in questions:
                    answer = feedback_response.answers.filter(question=question).first()
                    if answer:
                        if question.question_type == 'multiple_choice':
                            row.append(answer.selected_option or 'N/A')
                        elif question.question_type == 'rating':
                            row.append(str(answer.rating_value) if answer.rating_value else 'N/A')
                        else:  # text, textarea
                            row.append(answer.answer_text or 'N/A')
                    else:
                        row.append('N/A')

                writer.writerow(row)

            return response

        except Exception as e:
            return Response(
                {'error': f'Unable to export responses: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        """Export form responses to PDF"""
        try:
            form = self.get_object()
            responses = FeedbackResponse.objects.filter(form=form).order_by('-submitted_at')

            # Create PDF response
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{form.title}_responses.pdf"'

            # Create PDF document
            doc = SimpleDocTemplate(response, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph(f"Form Responses: {form.title}", title_style))
            story.append(Spacer(1, 20))

            # Form info
            info_style = styles['Normal']
            story.append(Paragraph(f"<b>Created by:</b> {form.created_by.username}", info_style))
            story.append(Paragraph(f"<b>Created at:</b> {form.created_at.strftime('%Y-%m-%d %H:%M:%S')}", info_style))
            story.append(Paragraph(f"<b>Total responses:</b> {responses.count()}", info_style))
            story.append(Spacer(1, 20))

            # Get all questions for this form
            questions = form.questions.all().order_by('order')

            if responses.exists():
                # Create table data
                table_data = []

                # Header row
                headers = ['Response ID', 'Submitted At', 'IP Address']
                for question in questions:
                    headers.append(f"{question.text[:30]}...")  # Truncate long questions
                table_data.append(headers)

                # Data rows
                for r in responses:
                    row = [
                        str(r.id)[:8],
                        r.submitted_at.strftime('%Y-%m-%d %H:%M'),
                        r.ip_address[:15] if r.ip_address else 'N/A'
                    ]

                    for q in questions:
                        ans = r.answers.filter(question=q).first()
                        if ans:
                            # Handle multiple types via JSON field or text
                            if ans.answer_value:
                                # Example: if stored like {"selected": "Option A"} or {"rating": 4}
                                if 'selected' in ans.answer_value:
                                    value = ans.answer_value['selected']
                                elif 'rating' in ans.answer_value:
                                    value = str(ans.answer_value['rating'])
                                else:
                                    # fallback for any other structured data
                                    value = str(ans.answer_value)
                            else:
                                # For plain text answers
                                txt = ans.answer_text or 'N/A'
                                value = txt[:30] + "..." if len(txt) > 30 else txt
                        else:
                            value = 'N/A'
                        row.append(value)

                    table_data.append(row)


                # Create table
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))

                story.append(table)
            else:
                story.append(Paragraph("No responses found for this form.", styles['Normal']))

            # Build PDF
            doc.build(story)
            return response

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                
                {'error': f'Unable to export responses: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


   

    @action(detail=True, methods=['get'])
    def export_analytics_excel(self, request, pk=None):
        """Export comprehensive analytics for a specific form to Excel"""
        try:
            form = self.get_object()

            # Create workbook with multiple sheets
            wb = openpyxl.Workbook()

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            subheader_font = Font(bold=True, color="000000")
            subheader_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

            # Sheet 1: Form Overview & Summary
            ws_overview = wb.active
            ws_overview.title = "Form Overview"

            # Form basic info
            analytics, created = FormAnalytics.objects.get_or_create(form=form)
            analytics.update_analytics()

            # Title
            title_cell = ws_overview.cell(row=1, column=1, value=f"Analytics Report: {form.title}")
            title_cell.font = Font(bold=True, size=16)
            ws_overview.merge_cells('A1:D1')

            # Basic stats
            info_cell = ws_overview.cell(row=3, column=1, value="Form Information")
            info_cell.font = subheader_font
            info_cell.fill = subheader_fill

            overview_data = [
                ['Form Title', form.title],
                ['Form Type', form.get_form_type_display()],
                ['Created Date', form.created_at.strftime('%Y-%m-%d %H:%M:%S')],
                ['Status', 'Active' if form.is_active else 'Inactive'],
                ['Total Questions', form.questions.count()],
                ['Total Responses', analytics.total_responses],
                ['Last Updated', analytics.last_updated.strftime('%Y-%m-%d %H:%M:%S')],
                ['Description', form.description or 'No description'],
            ]

            # Write overview data
            for row_num, (label, value) in enumerate(overview_data, 4):
                label_cell = ws_overview.cell(row=row_num, column=1, value=label)
                label_cell.font = Font(bold=True)
                ws_overview.cell(row=row_num, column=2, value=value)

            # Auto-adjust column widths
            ws_overview.column_dimensions['A'].width = 20
            ws_overview.column_dimensions['B'].width = 50

            # Add summary charts to overview sheet
            try:
                # Get summary data for charts
                questions = form.questions.all().order_by('order')

                # 1. Question Types Distribution Chart
                question_types = {}
                for question in questions:
                    q_type = question.get_question_type_display()
                    question_types[q_type] = question_types.get(q_type, 0) + 1

                if question_types:
                    # Add question types data
                    ws_overview.cell(row=15, column=1, value="Question Types Distribution").font = subheader_font
                    ws_overview.cell(row=15, column=1).fill = subheader_fill

                    current_row = 16
                    for q_type, count in question_types.items():
                        ws_overview.cell(row=current_row, column=1, value=q_type)
                        ws_overview.cell(row=current_row, column=2, value=count)
                        current_row += 1

                    # Create pie chart for question types
                    types_pie = PieChart()
                    types_pie.title = "Question Types Distribution"
                    types_pie.style = 26

                    types_labels = Reference(ws_overview, min_col=1, min_row=16, max_row=current_row-1)
                    types_data = Reference(ws_overview, min_col=2, min_row=16, max_row=current_row-1)

                    types_pie.add_data(types_data, titles_from_data=False)
                    types_pie.set_categories(types_labels)
                    types_pie.width = 12
                    types_pie.height = 10

                    ws_overview.add_chart(types_pie, "D15")

                # 2. Response Rate Chart (if there are responses)
                if analytics.total_responses > 0:
                    ws_overview.cell(row=current_row + 2, column=1, value="Response Summary").font = subheader_font
                    ws_overview.cell(row=current_row + 2, column=1).fill = subheader_fill

                    summary_row = current_row + 3
                    ws_overview.cell(row=summary_row, column=1, value="Total Responses")
                    ws_overview.cell(row=summary_row, column=2, value=analytics.total_responses)
                    ws_overview.cell(row=summary_row + 1, column=1, value="Total Questions")
                    ws_overview.cell(row=summary_row + 1, column=2, value=form.questions.count())

                    # Create column chart for summary
                    summary_chart = BarChart()
                    summary_chart.type = "col"
                    summary_chart.style = 10
                    summary_chart.title = "Form Summary"
                    summary_chart.y_axis.title = 'Count'

                    summary_labels = Reference(ws_overview, min_col=1, min_row=summary_row, max_row=summary_row+1)
                    summary_data = Reference(ws_overview, min_col=2, min_row=summary_row, max_row=summary_row+1)

                    summary_chart.add_data(summary_data, titles_from_data=False)
                    summary_chart.set_categories(summary_labels)
                    summary_chart.width = 12
                    summary_chart.height = 8

                    ws_overview.add_chart(summary_chart, f"D{summary_row}")

            except Exception as chart_error:
                # If chart creation fails, continue without charts
                pass

            # Sheet 2: Rating Questions Analytics
            questions = form.questions.all().order_by('order')
            rating_questions = [q for q in questions if q.question_type in ['rating', 'rating_10']]

            print(f"Form '{form.title}' has {questions.count()} total questions")
            print(f"Rating questions found: {len(rating_questions)}")
            for q in rating_questions:
                answer_count = Answer.objects.filter(question=q).count()
                print(f"  - {q.text[:50]}... has {answer_count} answers")

            if rating_questions:
                ws_ratings = wb.create_sheet(title="Rating Analytics")

                # Title
                title_cell = ws_ratings.cell(row=1, column=1, value="Rating Questions Analytics")
                title_cell.font = Font(bold=True, size=14)
                ws_ratings.merge_cells('A1:F1')

                # Add note about charts
                ws_ratings.cell(row=2, column=1, value="📊 Charts are positioned to the right of each question's data").font = Font(italic=True, color="0066CC")
                ws_ratings.merge_cells('A2:F2')

                current_row = 4
                for question in rating_questions:
                    try:
                        answers = Answer.objects.filter(question=question)
                        response_count = answers.count()

                        if response_count == 0:
                            continue

                        # Question header
                        question_cell = ws_ratings.cell(row=current_row, column=1, value=f"Q: {question.text}")
                        question_cell.font = subheader_font
                        question_cell.fill = subheader_fill
                        ws_ratings.merge_cells(f'A{current_row}:F{current_row}')
                        current_row += 1

                        # Calculate rating distribution
                        max_rating = 10 if question.question_type == 'rating_10' else 5
                        distribution = {}
                        total_rating_sum = 0
                        valid_ratings = 0

                        for i in range(1, max_rating + 1):
                            count = answers.filter(answer_text=str(i)).count()
                            distribution[str(i)] = count
                            total_rating_sum += i * count
                            valid_ratings += count

                        avg_rating = total_rating_sum / valid_ratings if valid_ratings > 0 else 0

                        # Summary stats
                        avg_cell = ws_ratings.cell(row=current_row, column=1, value="Average Rating:")
                        avg_cell.font = Font(bold=True)
                        ws_ratings.cell(row=current_row, column=2, value=f"{avg_rating:.1f}/{max_rating}")
                        total_cell = ws_ratings.cell(row=current_row, column=3, value="Total Responses:")
                        total_cell.font = Font(bold=True)
                        ws_ratings.cell(row=current_row, column=4, value=response_count)
                        current_row += 2

                        # Rating distribution headers
                        headers = ['Rating', 'Count', 'Percentage', 'Visual Bar']
                        for col_num, header in enumerate(headers, 1):
                            cell = ws_ratings.cell(row=current_row, column=col_num, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        current_row += 1

                        # Rating distribution data (highest to lowest)
                        chart_start_row = current_row
                        for i in range(max_rating, 0, -1):
                            count = distribution.get(str(i), 0)
                            percentage = (count / response_count * 100) if response_count > 0 else 0
                            bar_length = min(int(percentage / 5), 20)  # Scale for visual bar, max 20
                            visual_bar = '█' * bar_length + '░' * (20 - bar_length)

                            ws_ratings.cell(row=current_row, column=1, value=f"{i} Star{'s' if i != 1 else ''}")
                            ws_ratings.cell(row=current_row, column=2, value=count)
                            ws_ratings.cell(row=current_row, column=3, value=f"{percentage:.1f}%")
                            ws_ratings.cell(row=current_row, column=4, value=visual_bar)
                            current_row += 1

                        # Create EXACT horizontal bar chart matching your dashboard
                        try:
                            # Add average rating display (like dashboard shows 4.5 ⭐⭐⭐⭐⭐)
                            ws_ratings.cell(row=current_row, column=1, value="Average Rating:").font = Font(bold=True, size=14)
                            ws_ratings.cell(row=current_row, column=2, value=f"{avg_rating:.1f}").font = Font(bold=True, size=20, color="FF8C00")
                            ws_ratings.cell(row=current_row, column=3, value="⭐" * int(avg_rating)).font = Font(size=16, color="FF8C00")
                            ws_ratings.cell(row=current_row, column=4, value=f"({response_count} ratings)").font = Font(size=10, color="666666")
                            current_row += 2

                            # Create data for horizontal bars (EXACTLY like your dashboard)
                            chart_data_start = current_row

                            # Headers for the chart data
                            ws_ratings.cell(row=current_row, column=1, value="Rating").font = Font(bold=True)
                            ws_ratings.cell(row=current_row, column=2, value="Count").font = Font(bold=True)
                            ws_ratings.cell(row=current_row, column=3, value="Percentage").font = Font(bold=True)
                            current_row += 1

                            # Add data in descending order (5 stars at top, like dashboard)
                            for i in range(max_rating, 0, -1):
                                count = distribution.get(str(i), 0)
                                percentage = (count / response_count * 100) if response_count > 0 else 0

                                ws_ratings.cell(row=current_row, column=1, value=f"{i} ⭐")
                                ws_ratings.cell(row=current_row, column=2, value=count)
                                ws_ratings.cell(row=current_row, column=3, value=f"{percentage:.1f}%")
                                current_row += 1

                            # Create horizontal bar chart (EXACTLY like dashboard)
                            chart = BarChart()
                            chart.type = "bar"  # Horizontal bars
                            chart.style = 10
                            chart.title = question.text
                            chart.y_axis.title = None  # Remove axis titles for cleaner look
                            chart.x_axis.title = None

                            # Data range for chart
                            labels = Reference(ws_ratings, min_col=1, min_row=chart_data_start+1, max_row=current_row-1)
                            data = Reference(ws_ratings, min_col=2, min_row=chart_data_start+1, max_row=current_row-1)

                            chart.add_data(data, titles_from_data=False)
                            chart.set_categories(labels)

                            # Style to match dashboard (orange bars)
                            series = chart.series[0]
                            series.graphicalProperties.solidFill = "FF8C00"  # Orange color like dashboard

                            # Size and position
                            chart.width = 15
                            chart.height = 10

                            # Position chart to the right of data
                            chart_position = f"F{chart_data_start}"
                            ws_ratings.add_chart(chart, chart_position)

                            print(f"Added EXACT rating chart for '{question.text[:30]}' at {chart_position}")

                        except Exception as chart_error:
                            print(f"Chart creation error for rating question: {chart_error}")
                            pass

                        current_row += 35  # Space for multiple charts and between questions
                    except Exception as e:
                        # Skip this question if there's an error
                        continue

                # Auto-adjust column widths
                ws_ratings.column_dimensions['A'].width = 15
                ws_ratings.column_dimensions['B'].width = 10
                ws_ratings.column_dimensions['C'].width = 12
                ws_ratings.column_dimensions['D'].width = 25

            # Sheet 3: Multiple Choice Analytics
            choice_questions = [q for q in questions if q.question_type in ['radio', 'checkbox']]

            print(f"Multiple choice questions found: {len(choice_questions)}")
            for q in choice_questions:
                answer_count = Answer.objects.filter(question=q).count()
                print(f"  - {q.text[:50]}... has {answer_count} answers")

            if choice_questions:
                ws_choices = wb.create_sheet(title="Multiple Choice Analytics")

                # Title
                ws_choices.cell(row=1, column=1, value="Multiple Choice Questions Analytics").font = Font(bold=True, size=14)
                ws_choices.merge_cells('A1:E1')

                # Add note about charts
                ws_choices.cell(row=2, column=1, value="📊 Bar charts are positioned to the right of each question's data").font = Font(italic=True, color="0066CC")
                ws_choices.merge_cells('A2:E2')

                current_row = 4
                for question in choice_questions:
                    answers = Answer.objects.filter(question=question)
                    response_count = answers.count()

                    if response_count == 0:
                        continue

                    # Question header
                    ws_choices.cell(row=current_row, column=1, value=f"Q: {question.text}").font = subheader_font
                    ws_choices.cell(row=current_row, column=1).fill = subheader_fill
                    ws_choices.merge_cells(f'A{current_row}:E{current_row}')
                    current_row += 1

                    # Calculate distribution
                    if question.question_type == 'checkbox':
                        distribution = {}
                        total_selections = 0
                        for answer in answers:
                            options = [opt.strip() for opt in answer.answer_text.split(',') if opt.strip()]
                            for option in options:
                                distribution[option] = distribution.get(option, 0) + 1
                                total_selections += 1
                    else:
                        distribution_data = answers.values('answer_text').annotate(
                            count=Count('answer_text')
                        ).order_by('-count')
                        distribution = {item['answer_text']: item['count'] for item in distribution_data}
                        total_selections = response_count

                    # Summary
                    ws_choices.cell(row=current_row, column=1, value="Total Responses:").font = Font(bold=True)
                    ws_choices.cell(row=current_row, column=2, value=response_count)
                    if question.question_type == 'checkbox':
                        ws_choices.cell(row=current_row, column=3, value="Total Selections:").font = Font(bold=True)
                        ws_choices.cell(row=current_row, column=4, value=total_selections)
                    current_row += 2

                    # Distribution headers
                    headers = ['Option', 'Count', 'Percentage', 'Visual Bar']
                    for col_num, header in enumerate(headers, 1):
                        cell = ws_choices.cell(row=current_row, column=col_num, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    current_row += 1

                    # Distribution data (sorted by count)
                    sorted_distribution = sorted(distribution.items(), key=lambda x: x[1], reverse=True)
                    max_count = max(distribution.values()) if distribution else 1

                    chart_start_row = current_row
                    for option, count in sorted_distribution:
                        percentage = (count / total_selections * 100) if total_selections > 0 else 0
                        bar_length = int((count / max_count) * 20)  # Scale for visual bar
                        visual_bar = '█' * bar_length + '░' * (20 - bar_length)

                        ws_choices.cell(row=current_row, column=1, value=option)
                        ws_choices.cell(row=current_row, column=2, value=count)
                        ws_choices.cell(row=current_row, column=3, value=f"{percentage:.1f}%")
                        ws_choices.cell(row=current_row, column=4, value=visual_bar)
                        current_row += 1

                    # Create EXACT bar chart matching your dashboard
                    try:
                        # Add selection summary with colored indicators (like dashboard)
                        ws_choices.cell(row=current_row + 1, column=1, value="Selection Summary:").font = Font(bold=True, size=12)
                        current_row += 2

                        # Color palette matching dashboard
                        colors = ["3B82F6", "10B981", "F59E0B", "EF4444", "8B5CF6", "06B6D4", "84CC16", "F97316", "EC4899", "6366F1"]

                        # Add colored legend (like dashboard selection summary)
                        legend_start = current_row
                        for i, (option, count) in enumerate(sorted_distribution):
                            color = colors[i % len(colors)]
                            percentage = (count / total_selections * 100) if total_selections > 0 else 0

                            # Create colored indicator and text
                            ws_choices.cell(row=current_row, column=1, value="●").font = Font(color=color, size=16)
                            ws_choices.cell(row=current_row, column=2, value=option).font = Font(bold=True)
                            ws_choices.cell(row=current_row, column=3, value=f"{count} ({percentage:.1f}%)").font = Font(color="666666")
                            current_row += 1

                        # Create column chart (EXACTLY like dashboard)
                        col_chart = BarChart()
                        col_chart.type = "col"
                        col_chart.style = 12
                        col_chart.title = question.text
                        col_chart.y_axis.title = 'Number of Selections'
                        col_chart.x_axis.title = None  # Clean look like dashboard

                        # Data for chart
                        labels = Reference(ws_choices, min_col=1, min_row=chart_start_row, max_row=chart_start_row + len(sorted_distribution) - 1)
                        data = Reference(ws_choices, min_col=2, min_row=chart_start_row, max_row=chart_start_row + len(sorted_distribution) - 1)

                        col_chart.add_data(data, titles_from_data=False)
                        col_chart.set_categories(labels)

                        # Apply EXACT colors from dashboard to each bar
                        series = col_chart.series[0]
                        for i in range(len(sorted_distribution)):
                            if i < len(colors):
                                pt = series.dPt[i]
                                pt.graphicalProperties.solidFill = colors[i]

                        # Size and position like dashboard
                        col_chart.width = 14
                        col_chart.height = 10
                        ws_choices.add_chart(col_chart, f"F{chart_start_row}")

                        print(f"Added EXACT multi-choice chart for '{question.text[:30]}' with {len(sorted_distribution)} options")

                    except Exception as chart_error:
                        print(f"Multi-choice chart creation error: {chart_error}")
                        pass

                    current_row += 40  # Space for multiple charts and between questions

                # Auto-adjust column widths
                ws_choices.column_dimensions['A'].width = 25
                ws_choices.column_dimensions['B'].width = 10
                ws_choices.column_dimensions['C'].width = 12
                ws_choices.column_dimensions['D'].width = 25

            # Sheet 4: Yes/No Analytics
            yesno_questions = [q for q in questions if q.question_type == 'yes_no']

            print(f"Yes/No questions found: {len(yesno_questions)}")
            for q in yesno_questions:
                answer_count = Answer.objects.filter(question=q).count()
                print(f"  - {q.text[:50]}... has {answer_count} answers")

            if yesno_questions:
                ws_yesno = wb.create_sheet(title="Yes-No Analytics")

                # Title
                ws_yesno.cell(row=1, column=1, value="Yes/No Questions Analytics").font = Font(bold=True, size=14)
                ws_yesno.merge_cells('A1:E1')

                # Add note about charts
                ws_yesno.cell(row=2, column=1, value="🥧 Pie charts are positioned to the right of each question's data").font = Font(italic=True, color="0066CC")
                ws_yesno.merge_cells('A2:E2')

                current_row = 4
                for question in yesno_questions:
                    answers = Answer.objects.filter(question=question)
                    response_count = answers.count()

                    if response_count == 0:
                        continue

                    # Question header
                    ws_yesno.cell(row=current_row, column=1, value=f"Q: {question.text}").font = subheader_font
                    ws_yesno.cell(row=current_row, column=1).fill = subheader_fill
                    ws_yesno.merge_cells(f'A{current_row}:E{current_row}')
                    current_row += 1

                    # Calculate Yes/No distribution
                    yes_count = answers.filter(answer_text='Yes').count()
                    no_count = answers.filter(answer_text='No').count()
                    yes_percentage = (yes_count / response_count * 100) if response_count > 0 else 0
                    no_percentage = (no_count / response_count * 100) if response_count > 0 else 0

                    # Summary
                    ws_yesno.cell(row=current_row, column=1, value="Total Responses:").font = Font(bold=True)
                    ws_yesno.cell(row=current_row, column=2, value=response_count)
                    current_row += 2

                    # Results headers
                    headers = ['Answer', 'Count', 'Percentage', 'Visual Representation']
                    for col_num, header in enumerate(headers, 1):
                        cell = ws_yesno.cell(row=current_row, column=col_num, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    current_row += 1

                    # Yes result
                    chart_start_row = current_row
                    yes_bar = '█' * int(yes_percentage / 5) + '░' * (20 - int(yes_percentage / 5))
                    ws_yesno.cell(row=current_row, column=1, value="✓ Yes")
                    ws_yesno.cell(row=current_row, column=2, value=yes_count)
                    ws_yesno.cell(row=current_row, column=3, value=f"{yes_percentage:.1f}%")
                    ws_yesno.cell(row=current_row, column=4, value=yes_bar)
                    current_row += 1

                    # No result
                    no_bar = '█' * int(no_percentage / 5) + '░' * (20 - int(no_percentage / 5))
                    ws_yesno.cell(row=current_row, column=1, value="✗ No")
                    ws_yesno.cell(row=current_row, column=2, value=no_count)
                    ws_yesno.cell(row=current_row, column=3, value=f"{no_percentage:.1f}%")
                    ws_yesno.cell(row=current_row, column=4, value=no_bar)
                    current_row += 1

                    # Create EXACT pie chart matching your dashboard
                    try:
                        # Add summary boxes like dashboard (green for Yes, red for No)
                        ws_yesno.cell(row=current_row + 1, column=1, value="✓ Yes Responses").font = Font(bold=True, color="10B981")
                        ws_yesno.cell(row=current_row + 1, column=2, value=yes_count).font = Font(bold=True, size=16, color="10B981")
                        ws_yesno.cell(row=current_row + 1, column=3, value=f"({yes_percentage:.1f}%)").font = Font(color="10B981")

                        ws_yesno.cell(row=current_row + 2, column=1, value="✗ No Responses").font = Font(bold=True, color="EF4444")
                        ws_yesno.cell(row=current_row + 2, column=2, value=no_count).font = Font(bold=True, size=16, color="EF4444")
                        ws_yesno.cell(row=current_row + 2, column=3, value=f"({no_percentage:.1f}%)").font = Font(color="EF4444")

                        # Create pie chart data (EXACTLY like dashboard)
                        pie_data_start = current_row + 4
                        ws_yesno.cell(row=pie_data_start, column=1, value="Yes")
                        ws_yesno.cell(row=pie_data_start, column=2, value=yes_count)
                        ws_yesno.cell(row=pie_data_start + 1, column=1, value="No")
                        ws_yesno.cell(row=pie_data_start + 1, column=2, value=no_count)

                        # Create pie chart (EXACTLY like dashboard)
                        pie_chart = PieChart()
                        pie_chart.title = question.text
                        pie_chart.style = 26

                        # Data for pie chart
                        labels = Reference(ws_yesno, min_col=1, min_row=pie_data_start, max_row=pie_data_start + 1)
                        data = Reference(ws_yesno, min_col=2, min_row=pie_data_start, max_row=pie_data_start + 1)

                        pie_chart.add_data(data, titles_from_data=False)
                        pie_chart.set_categories(labels)

                        # Style with EXACT colors from dashboard (Green for Yes, Red for No)
                        series = pie_chart.series[0]
                        # Set colors: Green for Yes (first), Red for No (second)
                        pt1 = series.dPt[0]  # Yes
                        pt1.graphicalProperties.solidFill = "10B981"  # Green
                        pt2 = series.dPt[1]  # No
                        pt2.graphicalProperties.solidFill = "EF4444"  # Red

                        # Position and size like dashboard
                        pie_chart.width = 12
                        pie_chart.height = 12
                        ws_yesno.add_chart(pie_chart, f"F{chart_start_row}")

                        print(f"Added EXACT Yes/No pie chart for '{question.text[:30]}' - Yes: {yes_count}, No: {no_count}")

                    except Exception as chart_error:
                        print(f"Pie chart creation error: {chart_error}")
                        pass

                    current_row += 35  # Space for multiple charts and between questions

                # Auto-adjust column widths
                ws_yesno.column_dimensions['A'].width = 15
                ws_yesno.column_dimensions['B'].width = 10
                ws_yesno.column_dimensions['C'].width = 12
                ws_yesno.column_dimensions['D'].width = 25

            # Sheet 5: Text Responses
            text_questions = [q for q in questions if q.question_type in ['text', 'textarea', 'email', 'phone']]

            if text_questions:
                ws_text = wb.create_sheet(title="Text Responses")

                # Title
                ws_text.cell(row=1, column=1, value="Text Responses").font = Font(bold=True, size=14)
                ws_text.merge_cells('A1:C1')

                current_row = 3
                for question in text_questions:
                    answers = Answer.objects.filter(question=question).exclude(answer_text='')
                    response_count = answers.count()

                    if response_count == 0:
                        continue

                    # Question header
                    ws_text.cell(row=current_row, column=1, value=f"Q: {question.text}").font = subheader_font
                    ws_text.cell(row=current_row, column=1).fill = subheader_fill
                    ws_text.merge_cells(f'A{current_row}:C{current_row}')
                    current_row += 1

                    # Response count
                    ws_text.cell(row=current_row, column=1, value=f"Total Responses: {response_count}").font = Font(bold=True)
                    current_row += 2

                    # Headers
                    headers = ['Response #', 'Response Text', 'Submitted Date']
                    for col_num, header in enumerate(headers, 1):
                        cell = ws_text.cell(row=current_row, column=col_num, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    current_row += 1

                    # Responses
                    for idx, answer in enumerate(answers.order_by('-response__submitted_at'), 1):
                        ws_text.cell(row=current_row, column=1, value=idx)
                        ws_text.cell(row=current_row, column=2, value=answer.answer_text[:500] + ('...' if len(answer.answer_text) > 500 else ''))
                        ws_text.cell(row=current_row, column=3, value=answer.response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'))
                        current_row += 1

                    current_row += 2  # Space between questions

                # Auto-adjust column widths
                ws_text.column_dimensions['A'].width = 12
                ws_text.column_dimensions['B'].width = 60
                ws_text.column_dimensions['C'].width = 20

            # Create HTTP response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{form.title}_comprehensive_analytics.xlsx"'

            # Save workbook to response
            wb.save(response)
            return response

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Analytics export error: {error_details}")  # For debugging
            return Response(
                {'error': f'Unable to export form analytics: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def export_analytics_csv(self, request, pk=None):
        """Export comprehensive analytics for a specific form to CSV"""
        try:
            form = self.get_object()
            analytics, created = FormAnalytics.objects.get_or_create(form=form)
            analytics.update_analytics()

            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{form.title}_analytics.csv"'

            # Create CSV writer
            writer = csv.writer(response)

            # Write form overview section
            writer.writerow(['FORM ANALYTICS REPORT'])
            writer.writerow(['Form Title', form.title])
            writer.writerow(['Created By', form.created_by.username])
            writer.writerow(['Created At', form.created_at.strftime('%Y-%m-%d %H:%M:%S')])
            writer.writerow(['Total Responses', analytics.total_responses])
            writer.writerow(['Completion Rate', f"{analytics.completion_rate:.2f}%"])
            writer.writerow(['Average Rating', f"{analytics.average_rating:.2f}"])
            writer.writerow([])  # Empty row

            # Write question analytics section
            writer.writerow(['QUESTION ANALYTICS'])
            writer.writerow(['Question', 'Type', 'Response Count', 'Average Rating', 'Answer Distribution'])

            questions = form.questions.all().order_by('order')
            for question in questions:
                answers = Answer.objects.filter(question=question)
                response_count = answers.count()

                if question.question_type == 'rating':
                    avg_rating = answers.aggregate(avg=Avg('rating_value'))['avg'] or 0
                    writer.writerow([
                        question.text,
                        question.question_type,
                        response_count,
                        f"{avg_rating:.2f}",
                        'N/A'
                    ])
                elif question.question_type == 'multiple_choice':
                    # Calculate distribution
                    distribution = {}
                    for option in question.options:
                        count = answers.filter(selected_option=option).count()
                        if count > 0:
                            distribution[option] = count

                    dist_str = '; '.join([f"{k}: {v}" for k, v in distribution.items()])
                    writer.writerow([
                        question.text,
                        question.question_type,
                        response_count,
                        'N/A',
                        dist_str
                    ])
                else:  # text, textarea
                    writer.writerow([
                        question.text,
                        question.question_type,
                        response_count,
                        'N/A',
                        'Text responses (see detailed export)'
                    ])

            return response

        except Exception as e:
            return Response(
                {'error': f'Unable to export analytics: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def export_analytics_pdf(self, request, pk=None):
        """Export comprehensive analytics for a specific form to PDF"""
        try:
            form = self.get_object()
            analytics, created = FormAnalytics.objects.get_or_create(form=form)
            analytics.update_analytics()

            # Create PDF response
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{form.title}_analytics.pdf"'

            # Create PDF document
            doc = SimpleDocTemplate(response, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph(f"Analytics Report: {form.title}", title_style))
            story.append(Spacer(1, 20))

            # Form overview section
            overview_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                textColor=colors.darkblue
            )
            story.append(Paragraph("Form Overview", overview_style))

            overview_data = [
                ['Form Title', form.title],
                ['Created By', form.created_by.username],
                ['Created At', form.created_at.strftime('%Y-%m-%d %H:%M:%S')],
                ['Total Responses', str(analytics.total_responses)],
                ['Completion Rate', f"{analytics.completion_rate:.2f}%"],
                ['Average Rating', f"{analytics.average_rating:.2f}"]
            ]

            overview_table = Table(overview_data, colWidths=[2*inch, 3*inch])
            overview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            story.append(overview_table)
            story.append(Spacer(1, 20))

            # Question analytics section
            story.append(Paragraph("Question Analytics", overview_style))

            questions = form.questions.all().order_by('order')
            question_data = [['Question', 'Type', 'Responses', 'Avg Rating', 'Top Answer']]

            for question in questions:
                answers = Answer.objects.filter(question=question)
                response_count = answers.count()

                if question.question_type == 'rating':
                    avg_rating = answers.aggregate(avg=Avg('rating_value'))['avg'] or 0
                    top_answer = f"{avg_rating:.2f}"
                elif question.question_type == 'multiple_choice':
                    # Find most common answer
                    distribution = {}
                    for option in question.options:
                        count = answers.filter(selected_option=option).count()
                        if count > 0:
                            distribution[option] = count

                    if distribution:
                        top_answer = max(distribution, key=distribution.get)
                        top_answer = f"{top_answer} ({distribution[top_answer]})"
                    else:
                        top_answer = "No responses"
                    avg_rating = "N/A"
                else:  # text, textarea
                    avg_rating = "N/A"
                    top_answer = "Text responses"

                question_data.append([
                    question.text[:40] + "..." if len(question.text) > 40 else question.text,
                    question.question_type.replace('_', ' ').title(),
                    str(response_count),
                    str(avg_rating),
                    top_answer[:30] + "..." if len(str(top_answer)) > 30 else str(top_answer)
                ])

            question_table = Table(question_data, colWidths=[2.5*inch, 1*inch, 0.8*inch, 0.8*inch, 1.9*inch])
            question_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            story.append(question_table)

            # Build PDF
            doc.build(story)
            return response

        except Exception as e:
            return Response(
                {'error': f'Unable to export analytics: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        

    
    @action(detail=True, methods=['get'])
    def question_analytics(self, request, pk=None):
    
        try:
            form = self.get_object()

            # Get section filter from query params
            section_id = request.GET.get('section_id')

            # Build query for questions
            questions_query = Question.objects.filter(section__form=form)

            # Apply section filter if provided
            if section_id:
                questions_query = questions_query.filter(section_id=section_id)

            questions = questions_query.select_related('section')

            # Get all sections for the dropdown
            sections = Section.objects.filter(form=form).values('id', 'title', 'order')

            question_analytics = []
            for question in questions:
                answers = Answer.objects.filter(question=question)
                response_count = answers.count()

                # Safe options handling
                options_list = []
                if question.options is not None:
                    if isinstance(question.options, list):
                        options_list = question.options
                    else:
                        try:
                            options_list = list(question.options) if hasattr(question.options, '__iter__') else []
                        except:
                            options_list = []

                analytics_data = {
                    "question_id": question.id,
                    "question_text": question.text,
                    "question_type": question.question_type,
                    "response_count": response_count,
                    "average_rating": None,
                    "answer_distribution": {},
                    "options": options_list,
                    "section_id": question.section.id,
                    "section_title": question.section.title,
                }

                # Handle analytics by question type
                if question.question_type == "rating":
                    total = 0
                    for answer in answers:
                        try:
                            total += int(answer.answer_text)
                        except ValueError:
                            pass
                    analytics_data["average_rating"] = round(total / response_count, 2) if response_count else None

                elif question.question_type in ["radio", "checkbox", "dropdown"]:  # ✅ Dropdown included here
                    if question.question_type == "checkbox":
                        distribution = {}
                        for answer in answers:
                            options = [opt.strip() for opt in answer.answer_text.split(",") if opt.strip()]
                            for option in options:
                                distribution[option] = distribution.get(option, 0) + 1
                        analytics_data["answer_distribution"] = distribution

                    elif question.question_type == "dropdown":
                        # ✅ Dropdown logic - single selection
                        distribution = {opt: 0 for opt in options_list}
                        for answer in answers:
                            answer_text = answer.answer_text.strip()
                            if answer_text in distribution:
                                distribution[answer_text] += 1
                            else:
                                distribution[answer_text] = distribution.get(answer_text, 0) + 1
                        analytics_data["answer_distribution"] = distribution

                    else:  # Radio button logic
                        distribution = {opt: 0 for opt in options_list}
                        for answer in answers:
                            answer_text = answer.answer_text.strip()
                            if answer_text in distribution:
                                distribution[answer_text] += 1
                            else:
                                distribution[answer_text] = distribution.get(answer_text, 0) + 1
                        analytics_data["answer_distribution"] = distribution

                # Add this question's analytics
                question_analytics.append(analytics_data)

            return Response({
                "sections": list(sections),
                "questions": question_analytics,
                "selected_section": section_id
            }, status=status.HTTP_200_OK)

        except Exception as e:
            print(f"Analytics error: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {"error": f"Unable to load question analytics: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class PublicFormsListView(APIView):
    """Public view for listing all active feedback forms"""
    permission_classes = [permissions.AllowAny]
    
    def get(self, request):
        """Get list of all active public forms"""
        forms = FeedbackForm.objects.filter(
            is_active=True
        ).order_by('-created_at')
        
        # Filter out expired forms
        active_forms = [form for form in forms if not form.is_expired]
        
        serializer = FeedbackFormSerializer(active_forms, many=True)
        return Response(serializer.data)


class PublicFeedbackFormView(APIView):
    """Public view for accessing feedback forms via shareable links"""
    permission_classes = [permissions.AllowAny]
    
    def get(self, request, form_id):
        """Get form details for public access"""
        try:
            form = FeedbackForm.objects.get(id=form_id, is_active=True)
            
            # Check if form is expired
            if form.is_expired:
                return Response(
                    {'error': 'This form has expired'}, 
                    status=status.HTTP_410_GONE
                )
            
            # DEBUG: Check database state
            print("🔍 DATABASE DEBUG - Form:", form.title)
            print("🔍 DATABASE DEBUG - Sections:", form.sections.count())
            for section in form.sections.all():
                print(f"🔍 DATABASE DEBUG - Section {section.id}: '{section.title}'")
                for question in section.questions.all():
                    print(f"🔍 DATABASE DEBUG - Question {question.id}: '{question.text}'")
            
            serializer = FeedbackFormSerializer(form)
            
            # DEBUG: Check serialized data
            print("🔍 SERIALIZED DEBUG - Questions:", len(serializer.data.get('questions', [])))
            for q in serializer.data.get('questions', []):
                print(f"🔍 SERIALIZED DEBUG - Q{q['id']}: '{q['text']}'")
            
            return Response(serializer.data)
        
        except FeedbackForm.DoesNotExist:
            return Response(
                {'error': 'Form not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
    
    def post(self, request, form_id):
        """Submit a feedback response"""
        try:
            form = FeedbackForm.objects.get(id=form_id, is_active=True)
            
            # Check if form is expired
            if form.is_expired:
                return Response(
                    {'error': 'This form has expired'}, 
                    status=status.HTTP_410_GONE
                )
            
            # 🔥 FIX: Get all required questions from all sections
            required_questions = []
            form_question_ids = []
            
            for section in form.sections.all():
                section_required = section.questions.filter(is_required=True)
                required_questions.extend(section_required)
                
                # Also collect all question IDs for validation
                section_questions = section.questions.all()
                form_question_ids.extend([q.id for q in section_questions])
            
            print(f"🔍 DEBUG: Found {len(required_questions)} required questions")
            print(f"🔍 DEBUG: Form has {len(form_question_ids)} total questions")
            
            submitted_answers = request.data.get('answers', [])
            submitted_question_ids = [answer['question'] for answer in submitted_answers]
            
            print(f"🔍 DEBUG: Received {len(submitted_answers)} answers")
            print(f"🔍 DEBUG: Submitted question IDs: {submitted_question_ids}")
            
            # Validate that all required questions are answered
            missing_required = []
            for question in required_questions:
                if question.id not in submitted_question_ids:
                    missing_required.append({
                        'question_id': question.id,
                        'question_text': question.text
                    })
                    print(f"❌ MISSING REQUIRED: Q{question.id} - '{question.text}'")
            
            if missing_required:
                return Response(
                    {
                        'error': 'Missing required questions',
                        'missing_questions': missing_required
                    }, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate that all submitted questions exist in this form
            invalid_questions = []
            for answer in submitted_answers:
                question_id = answer.get('question')
                if question_id not in form_question_ids:
                    invalid_questions.append(question_id)
                    print(f"❌ INVALID QUESTION: {question_id}")
            
            if invalid_questions:
                return Response(
                    {
                        'error': 'Invalid questions submitted',
                        'invalid_questions': invalid_questions
                    }, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create response
            response_data = {
                'form': form.id,
                'answers': submitted_answers
            }
            
            serializer = FeedbackResponseCreateSerializer(
                data=response_data, 
                context={'request': request}
            )
            
            if serializer.is_valid():
                response = serializer.save()
                
                # Update analytics
                analytics, created = FormAnalytics.objects.get_or_create(form=form)
                analytics.update_analytics()
                
                # Create notification record in database
                Notification.objects.create(
                    user=form.created_by,
                    notification_type='new_response',
                    title='New Response Received',
                    message=f'New response received for "{form.title}"',
                    data={
                        "form_id": str(form.id),
                        "response_id": str(response.id),
                        "form_title": form.title
                    }
                )
                
                # Send real-time notification to form creator
                send_notification_to_group(
                    f"user_{form.created_by.id}",
                    "new_response",
                    f"New response received for '{form.title}'",
                    {
                        "form_id": str(form.id),
                        "response_id": str(response.id),
                        "form_title": form.title
                    }
                )
                
                print(f"✅ SUCCESS: Response {response.id} submitted successfully")
                return Response({
                    'message': 'Feedback submitted successfully',
                    'response_id': str(response.id)
                }, status=status.HTTP_201_CREATED)
            
            print(f"❌ SERIALIZER ERRORS: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        except FeedbackForm.DoesNotExist:
            print(f"❌ FORM NOT FOUND: {form_id}")
            return Response(
                {'error': 'Form not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            print(f"💥 ERROR in PublicFeedbackFormView: {str(e)}")
            import traceback
            print(f"💥 TRACEBACK: {traceback.format_exc()}")
            return Response(
                {'error': 'Internal server error'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
class FeedbackResponseViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing feedback responses"""
    serializer_class = FeedbackResponseSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user

        # If superuser, show all responses and allow filtering by admin_name
        # if user.is_superuser:
        #     queryset = FeedbackResponse.objects.all()


        #     admin_name = self.request.query_params.get('admin_name')
        #     if admin_name:
        #         queryset = queryset.filter(created_by__username=admin_name)
        #     return queryset


        return FeedbackResponse.objects.filter(form__created_by=user)


    # def create(self, request, *args, **kwargs):
    #     try:
    #         with transaction.atomic():
    #             return super().create(request, *args, **kwargs)
    #     except Exception as e:
    #         return Response(
    #             {"error": f"Failed to submit response: {str(e)}"}, 
    #             status=status.HTTP_400_BAD_REQUEST
    #         )    

    @action(detail=False, methods=['get'])
    def export_all_excel(self, request):
        """Export all responses from all forms to Excel"""
        try:
            user = request.user
            if user.is_superuser:
                forms = FeedbackForm.objects.all()
            else:
                forms = FeedbackForm.objects.filter(created_by=user)

            # responses = self.get_queryset().order_by('-submitted_at')
            responses = FeedbackResponse.objects.filter(form__in=forms).order_by('-submitted_at')

            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "All Responses"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Create headers
            headers = ['Response ID', 'Form Title', 'Submitted At', 'IP Address', 'Question', 'Question Type', 'Answer']

            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Write data
            row_num = 2
            for response in responses:
                for answer in response.answers.all():
                    ws.cell(row=row_num, column=1, value=str(response.id))
                    ws.cell(row=row_num, column=2, value=response.form.title)
                    ws.cell(row=row_num, column=3, value=response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row_num, column=4, value=response.ip_address or 'N/A')
                    ws.cell(row=row_num, column=5, value=answer.question.text)
                    ws.cell(row=row_num, column=6, value=answer.question.question_type)
                    ws.cell(row=row_num, column=7, value=answer.answer_text)
                    row_num += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Create HTTP response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="all_responses_{user.username}.xlsx"'

            # Save workbook to response
            wb.save(response)
            return response

        except Exception as e:
            return Response(
                {'error': f'Unable to export all responses: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def export_all_csv(self, request):
        """Export all responses from all forms to CSV"""
        try:
            user = request.user
            if user.is_superuser:
                forms = FeedbackForm.objects.all()
            else:
                forms = FeedbackForm.objects.filter(created_by=user)

            # responses = self.get_queryset().order_by('-submitted_at')
            responses = FeedbackResponse.objects.filter(form__in=forms).order_by('-submitted_at')

            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="all_responses_{user.username}.csv"'

            # Create CSV writer
            writer = csv.writer(response)

            # Create headers
            headers = ['Response ID', 'Form Title', 'Submitted At', 'IP Address', 'Question', 'Question Type', 'Answer']
            writer.writerow(headers)

            # Write data rows
            for feedback_response in responses:
                for answer in feedback_response.answers.all():
                    question = answer.question

                    if question.question_type == 'multiple_choice':
                        answer_text = answer.selected_option or 'N/A'
                    elif question.question_type == 'rating':
                        answer_text = str(answer.answer_value.get('value')) if answer.answer_value.get('value') else answer.answer_text or 'N/A'
                    else:  # text, textarea
                        answer_text = answer.answer_text or 'N/A'

                    writer.writerow([
                        str(feedback_response.id),
                        feedback_response.form.title,
                        feedback_response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
                        feedback_response.ip_address or 'N/A',
                        question.text,
                        question.question_type,
                        answer_text
                    ])

            return response

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Unable to export all responses: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def export_all_pdf(self, request):
        """Export all responses from all forms to PDF"""
        try:
            user = request.user
            if user.is_superuser:
                forms = FeedbackForm.objects.all()
            else:
                forms = FeedbackForm.objects.filter(created_by=user)

            # responses = self.get_queryset().order_by('-submitted_at')
            responses = FeedbackResponse.objects.filter(form__in=forms).order_by('-submitted_at')
            

            # Create PDF response
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="all_responses_{user.username}.pdf"'

            # Create PDF document
            doc = SimpleDocTemplate(response, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph(f"All Responses Report - {user.username}", title_style))
            story.append(Spacer(1, 20))

            # Summary info
            info_style = styles['Normal']
            story.append(Paragraph(f"<b>Generated:</b> {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", info_style))
            story.append(Paragraph(f"<b>Total responses:</b> {responses.count()}", info_style))
            story.append(Spacer(1, 20))

            if responses.exists():
                # Create table data
                table_data = []

                # Header row
                headers = ['Response ID', 'Form Title', 'Submitted At', 'Question', 'Answer']
                table_data.append(headers)

                # Data rows
                for feedback_response in responses:
                    for answer in feedback_response.answers.all():
                        question = answer.question

                        if question.question_type == 'multiple_choice':
                            answer_text = answer.selected_option or 'N/A'
                        elif question.question_type == 'rating':
                            # answer_text = str(answer.rating_value) if answer.rating_value else 'N/A'
                            answer_text = str(answer.answer_value) if answer.answer_value else answer.answer_text or 'N/A'

                        else:  # text, textarea
                            answer_text = (answer.answer_text[:50] + '...') if answer.answer_text and len(answer.answer_text) > 50 else (answer.answer_text or 'N/A')

                        table_data.append([
                            str(feedback_response.id)[:8],  # Truncate ID
                            feedback_response.form.title[:20] + '...' if len(feedback_response.form.title) > 20 else feedback_response.form.title,
                            feedback_response.submitted_at.strftime('%Y-%m-%d %H:%M'),
                            question.text[:30] + '...' if len(question.text) > 30 else question.text,
                            answer_text
                        ])

                # Create table
                table = Table(table_data, colWidths=[1*inch, 2*inch, 1.5*inch, 2*inch, 2.5*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))

                story.append(table)
            else:
                story.append(Paragraph("No responses found.", styles['Normal']))

            # Build PDF
            doc.build(story)
            return response

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Unable to export all responses: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def export_analytics_excel(self, request):
        """Export comprehensive analytics for all forms to Excel"""
        try:
            user = request.user
            forms = FeedbackForm.objects.all() if user.is_superuser else FeedbackForm.objects.filter(created_by=user)


            # Create workbook with multiple sheets
            wb = openpyxl.Workbook()

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            subheader_font = Font(bold=True, color="000000")
            subheader_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

            # Sheet 1: Forms Overview & Summary
            ws_overview = wb.active
            ws_overview.title = "Forms Overview"

            # Title
            ws_overview.cell(row=1, column=1, value=f"Comprehensive Analytics Report - {user.username}").font = Font(bold=True, size=16)
            ws_overview.merge_cells('A1:I1')

            # Add note about charts
            ws_overview.cell(row=2, column=1, value="📊 Charts are embedded in this Excel file. Scroll right or check other sheets to view visual analytics.").font = Font(italic=True, color="0066CC")
            ws_overview.merge_cells('A2:I2')

            # Summary stats
            total_forms = forms.count()
            total_responses = sum(form.responses.count() for form in forms)
            active_forms = forms.filter(is_active=True).count()

            ws_overview.cell(row=4, column=1, value="Summary Statistics").font = subheader_font
            ws_overview.cell(row=4, column=1).fill = subheader_fill
            ws_overview.merge_cells('A4:C4')

            summary_data = [
                ['Total Forms', total_forms],
                ['Active Forms', active_forms],
                ['Total Responses', total_responses],
                ['Report Generated', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
            ]

            for row_num, (label, value) in enumerate(summary_data, 5):
                ws_overview.cell(row=row_num, column=1, value=label).font = Font(bold=True)
                ws_overview.cell(row=row_num, column=2, value=value)

            # Create a simple test chart with the summary data
            try:
                test_chart = BarChart()
                test_chart.type = "col"
                test_chart.style = 10
                test_chart.title = "Summary Statistics"
                test_chart.y_axis.title = 'Count'
                test_chart.x_axis.title = 'Metrics'

                # Use first 3 rows of summary data (exclude date)
                test_labels = Reference(ws_overview, min_col=1, min_row=5, max_row=7)
                test_data = Reference(ws_overview, min_col=2, min_row=5, max_row=7)

                test_chart.add_data(test_data, titles_from_data=False)
                test_chart.set_categories(test_labels)

                # Position chart in a very visible location
                test_chart.width = 10
                test_chart.height = 6
                ws_overview.add_chart(test_chart, "D5")  # Right next to the data

                print(f"Added test chart at D5")

            except Exception as chart_error:
                print(f"Test chart error: {chart_error}")
                pass

            # Forms details
            ws_overview.cell(row=9, column=1, value="Forms Details").font = subheader_font
            ws_overview.cell(row=9, column=1).fill = subheader_fill

            overview_headers = [
                'Form Title', 'Form Type', 'Created Date', 'Status', 'Total Questions',
                'Total Responses', 'Last Response Date'
            ]

            # Write overview headers
            for col_num, header in enumerate(overview_headers, 1):
                cell = ws_overview.cell(row=10, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Write overview data
            for row_num, form in enumerate(forms, 11):
                analytics, created = FormAnalytics.objects.get_or_create(form=form)
                analytics.update_analytics()

                last_response = form.responses.order_by('-submitted_at').first()
                last_response_date = last_response.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if last_response else 'No responses'

                ws_overview.cell(row=row_num, column=1, value=form.title)
                ws_overview.cell(row=row_num, column=2, value=form.get_form_type_display())
                ws_overview.cell(row=row_num, column=3, value=form.created_at.strftime('%Y-%m-%d'))
                ws_overview.cell(row=row_num, column=4, value='Active' if form.is_active else 'Inactive')
                ws_overview.cell(row=row_num, column=5, value=form.questions.count())
                ws_overview.cell(row=row_num, column=6, value=analytics.total_responses)
                ws_overview.cell(row=row_num, column=7, value=last_response_date)

            # Add a simple visible chart to the overview
            try:
                if forms.count() > 0:
                    # Create a chart showing response counts
                    overview_chart = BarChart()
                    overview_chart.type = "col"
                    overview_chart.style = 10
                    overview_chart.title = "Response Count by Form"
                    overview_chart.y_axis.title = 'Number of Responses'
                    overview_chart.x_axis.title = 'Forms'

                    # Use form titles and response counts
                    chart_labels = Reference(ws_overview, min_col=1, min_row=11, max_row=10+forms.count())
                    chart_data = Reference(ws_overview, min_col=6, min_row=11, max_row=10+forms.count())

                    overview_chart.add_data(chart_data, titles_from_data=False)
                    overview_chart.set_categories(chart_labels)

                    # Position chart in a visible location
                    overview_chart.width = 12
                    overview_chart.height = 8
                    ws_overview.add_chart(overview_chart, "I11")  # Right side, visible

                    print(f"Added overview chart at I11 with {forms.count()} forms")

            except Exception as chart_error:
                print(f"Overview chart error: {chart_error}")
                pass

            # Auto-adjust column widths for overview
            for column in ws_overview.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_overview.column_dimensions[column_letter].width = adjusted_width

            # Sheet 2: Rating Questions Summary
            all_rating_questions = []
            for form in forms:
                questions = form.questions.filter(question_type__in=['rating', 'rating_10'])
                for question in questions:
                    if Answer.objects.filter(question=question).count() > 0:
                        all_rating_questions.append((form, question))

            if all_rating_questions:
                ws_ratings = wb.create_sheet(title="Rating Questions Summary")

                # Title
                ws_ratings.cell(row=1, column=1, value="Rating Questions Summary - All Forms").font = Font(bold=True, size=14)
                ws_ratings.merge_cells('A1:F1')

                # Headers
                headers = ['Form', 'Question', 'Type', 'Avg Rating', 'Total Responses', 'Rating Distribution']
                for col_num, header in enumerate(headers, 1):
                    cell = ws_ratings.cell(row=3, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                current_row = 4
                for form, question in all_rating_questions:
                    answers = Answer.objects.filter(question=question)
                    response_count = answers.count()

                    # Calculate average rating
                    max_rating = 10 if question.question_type == 'rating_10' else 5
                    total_rating_sum = 0
                    valid_ratings = 0

                    for i in range(1, max_rating + 1):
                        count = answers.filter(answer_text=str(i)).count()
                        total_rating_sum += i * count
                        valid_ratings += count

                    avg_rating = total_rating_sum / valid_ratings if valid_ratings > 0 else 0

                    # Create distribution summary
                    distribution_summary = []
                    for i in range(max_rating, 0, -1):
                        count = answers.filter(answer_text=str(i)).count()
                        if count > 0:
                            distribution_summary.append(f"{i}★:{count}")

                    ws_ratings.cell(row=current_row, column=1, value=form.title[:30] + ('...' if len(form.title) > 30 else ''))
                    ws_ratings.cell(row=current_row, column=2, value=question.text[:40] + ('...' if len(question.text) > 40 else ''))
                    ws_ratings.cell(row=current_row, column=3, value=f"{max_rating}-Point Scale")
                    ws_ratings.cell(row=current_row, column=4, value=f"{avg_rating:.1f}/{max_rating}")
                    ws_ratings.cell(row=current_row, column=5, value=response_count)
                    ws_ratings.cell(row=current_row, column=6, value=" | ".join(distribution_summary))
                    current_row += 1

                # Create summary chart for all rating questions
                try:
                    if current_row > 4:  # Only create chart if there's data
                        summary_chart = BarChart()
                        summary_chart.type = "col"
                        summary_chart.style = 10
                        summary_chart.title = "Average Ratings Across All Forms"
                        summary_chart.y_axis.title = 'Average Rating'
                        summary_chart.x_axis.title = 'Questions'

                        # Use form names and average ratings
                        labels = Reference(ws_ratings, min_col=1, min_row=4, max_row=current_row-1)
                        data = Reference(ws_ratings, min_col=4, min_row=4, max_row=current_row-1)

                        summary_chart.add_data(data, titles_from_data=False)
                        summary_chart.set_categories(labels)

                        # Position chart in a more visible location
                        summary_chart.width = 15
                        summary_chart.height = 10
                        ws_ratings.add_chart(summary_chart, "H4")  # Fixed position for visibility

                        print(f"Added rating summary chart at H4 with {current_row-4} data points")

                except Exception as chart_error:
                    print(f"Rating chart error: {chart_error}")
                    pass

                # Auto-adjust column widths
                ws_ratings.column_dimensions['A'].width = 35
                ws_ratings.column_dimensions['B'].width = 45
                ws_ratings.column_dimensions['C'].width = 15
                ws_ratings.column_dimensions['D'].width = 12
                ws_ratings.column_dimensions['E'].width = 15
                ws_ratings.column_dimensions['F'].width = 40

            # Sheet 3: Choice Questions Summary
            all_choice_questions = []
            for form in forms:
                questions = form.questions.filter(question_type__in=['radio', 'checkbox', 'yes_no'])
                for question in questions:
                    if Answer.objects.filter(question=question).count() > 0:
                        all_choice_questions.append((form, question))

            if all_choice_questions:
                ws_choices = wb.create_sheet(title="Choice Questions Summary")

                # Title
                ws_choices.cell(row=1, column=1, value="Choice Questions Summary - All Forms").font = Font(bold=True, size=14)
                ws_choices.merge_cells('A1:F1')

                # Headers
                headers = ['Form', 'Question', 'Type', 'Total Responses', 'Top Answer', 'Answer Distribution']
                for col_num, header in enumerate(headers, 1):
                    cell = ws_choices.cell(row=3, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                current_row = 4
                for form, question in all_choice_questions:
                    answers = Answer.objects.filter(question=question)
                    response_count = answers.count()

                    # Calculate distribution
                    if question.question_type == 'checkbox':
                        distribution = {}
                        for answer in answers:
                            options = [opt.strip() for opt in answer.answer_text.split(',') if opt.strip()]
                            for option in options:
                                distribution[option] = distribution.get(option, 0) + 1
                    else:
                        distribution_data = answers.values('answer_text').annotate(
                            count=Count('answer_text')
                        ).order_by('-count')
                        distribution = {item['answer_text']: item['count'] for item in distribution_data}

                    # Get top answer
                    top_answer = max(distribution, key=distribution.get) if distribution else 'N/A'
                    top_count = distribution.get(top_answer, 0)

                    # Create distribution summary
                    distribution_summary = []
                    sorted_dist = sorted(distribution.items(), key=lambda x: x[1], reverse=True)[:3]
                    for option, count in sorted_dist:
                        percentage = (count / response_count * 100) if response_count > 0 else 0
                        distribution_summary.append(f"{option}:{count}({percentage:.0f}%)")

                    ws_choices.cell(row=current_row, column=1, value=form.title[:30] + ('...' if len(form.title) > 30 else ''))
                    ws_choices.cell(row=current_row, column=2, value=question.text[:40] + ('...' if len(question.text) > 40 else ''))
                    ws_choices.cell(row=current_row, column=3, value=question.get_question_type_display())
                    ws_choices.cell(row=current_row, column=4, value=response_count)
                    ws_choices.cell(row=current_row, column=5, value=f"{top_answer} ({top_count})")
                    ws_choices.cell(row=current_row, column=6, value=" | ".join(distribution_summary))
                    current_row += 1

                # Create summary chart for choice questions
                try:
                    if current_row > 4:  # Only create chart if there's data
                        choice_chart = BarChart()
                        choice_chart.type = "col"
                        choice_chart.style = 12
                        choice_chart.title = "Response Counts Across Choice Questions"
                        choice_chart.y_axis.title = 'Number of Responses'
                        choice_chart.x_axis.title = 'Questions'

                        # Use form names and response counts
                        labels = Reference(ws_choices, min_col=1, min_row=4, max_row=current_row-1)
                        data = Reference(ws_choices, min_col=4, min_row=4, max_row=current_row-1)

                        choice_chart.add_data(data, titles_from_data=False)
                        choice_chart.set_categories(labels)

                        # Position chart below the data
                        choice_chart.width = 20
                        choice_chart.height = 12
                        ws_choices.add_chart(choice_chart, f"A{current_row + 2}")

                except Exception as chart_error:
                    pass

                # Auto-adjust column widths
                ws_choices.column_dimensions['A'].width = 35
                ws_choices.column_dimensions['B'].width = 45
                ws_choices.column_dimensions['C'].width = 15
                ws_choices.column_dimensions['D'].width = 15
                ws_choices.column_dimensions['E'].width = 25
                ws_choices.column_dimensions['F'].width = 50

            # Create HTTP response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="comprehensive_analytics_{user.username}_{timezone.now().strftime("%Y%m%d")}.xlsx"'

            # Save workbook to response
            wb.save(response)
            return response

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"All forms analytics export error: {error_details}")  # For debugging
            return Response(
                {'error': f'Unable to export analytics: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def export_analytics_csv(self, request):
        """Export comprehensive analytics for all forms to CSV"""
        try:
            user = request.user
            forms = FeedbackForm.objects.all() if user.is_superuser else FeedbackForm.objects.filter(created_by=user)

                        # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="comprehensive_analytics_{user.username}_{timezone.now().strftime("%Y%m%d")}.csv"'

            # Create CSV writer
            writer = csv.writer(response)

            # Write header
            writer.writerow(['COMPREHENSIVE ANALYTICS REPORT'])
            writer.writerow(['Generated', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
            writer.writerow(['User', user.username])
            writer.writerow([])  # Empty row

            # Forms overview section
            writer.writerow(['FORMS OVERVIEW'])
            writer.writerow(['Form Title', 'Status', 'Created', 'Total Responses', 'Completion Rate', 'Average Rating'])

            for form in forms:
                analytics, created = FormAnalytics.objects.get_or_create(form=form)
                analytics.update_analytics()

                writer.writerow([
                    form.title,
                    'Active' if form.is_active else 'Inactive',
                    form.created_at.strftime('%Y-%m-%d'),
                    analytics.total_responses,
                    f"{analytics.completion_rate:.2f}%",
                    f"{analytics.average_rating:.2f}"
                ])

            writer.writerow([])  # Empty row

            # Question analytics section
            writer.writerow(['QUESTION ANALYTICS'])
            writer.writerow(['Form', 'Question', 'Type', 'Responses', 'Average Rating', 'Top Answer'])

            for form in forms:
                questions = form.questions.all().order_by('order')
                for question in questions:
                    answers = Answer.objects.filter(question=question)
                    response_count = answers.count()

                    if question.question_type == 'rating':
                        # avg_rating = round(answers.aggregate(avg=(Avg('answer_value')))['avg'] or 0)
                        avg_rating = answers.annotate(
                                      rating=Cast(KeyTransform('rating', 'answer_value'), FloatField())
                        ).aggregate(avg=Avg('rating'))['avg'] or 0

                        top_answer = f"{avg_rating:.2f}"
                    elif question.question_type == 'multiple_choice':
                        # Find most common answer
                        distribution = {}
                        for option in question.options:
                            count = answers.filter(selected_option=option).count()
                            if count > 0:
                                distribution[option] = count

                        if distribution:
                            top_answer = max(distribution, key=distribution.get)
                            top_answer = f"{top_answer} ({distribution[top_answer]})"
                        else:
                            top_answer = "No responses"
                        avg_rating = "N/A"
                    else:  # text, textarea
                        avg_rating = "N/A"
                        top_answer = "Text responses"

                    writer.writerow([
                        form.title[:30] + '...' if len(form.title) > 30 else form.title,
                        question.text[:40] + '...' if len(question.text) > 40 else question.text,
                        question.question_type.replace('_', ' ').title(),
                        response_count,
                        str(avg_rating),
                        top_answer[:50] + '...' if len(str(top_answer)) > 50 else str(top_answer)
                    ])

            return response

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Unable to export analytics: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def export_analytics_pdf(self, request):
        """Export comprehensive analytics for all forms to PDF"""
        try:
            user = request.user
            forms = FeedbackForm.objects.all() if user.is_superuser else FeedbackForm.objects.filter(created_by=user)

            # Create PDF response
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="comprehensive_analytics_{user.username}_{timezone.now().strftime("%Y%m%d")}.pdf"'

            # Create PDF document
            doc = SimpleDocTemplate(response, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph(f"Comprehensive Analytics Report", title_style))
            story.append(Spacer(1, 20))

            # Report info
            info_style = styles['Normal']
            story.append(Paragraph(f"<b>Generated:</b> {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", info_style))
            story.append(Paragraph(f"<b>User:</b> {user.username}", info_style))
            story.append(Paragraph(f"<b>Total Forms:</b> {forms.count()}", info_style))
            story.append(Spacer(1, 20))

            # Forms overview section
            overview_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                textColor=colors.darkblue
            )
            story.append(Paragraph("Forms Overview", overview_style))

            # Forms overview table
            forms_data = [['Form Title', 'Status', 'Created', 'Responses', 'Completion Rate']]

            for form in forms:
                analytics, created = FormAnalytics.objects.get_or_create(form=form)
                analytics.update_analytics()

                forms_data.append([
                    form.title[:25] + '...' if len(form.title) > 25 else form.title,
                    'Active' if form.is_active else 'Inactive',
                    form.created_at.strftime('%Y-%m-%d'),
                    str(analytics.total_responses),
                    f"{analytics.completion_rate:.1f}%"
                ])

            forms_table = Table(forms_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch, 1.5*inch])
            forms_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            story.append(forms_table)
            story.append(Spacer(1, 20))
            
            
                    # Chart Data
            total_forms = forms.count()
            active_forms = forms.filter(is_active=True).count()
            total_responses = sum(form.responses.count() for form in forms)

            chart_data = [[total_forms, active_forms, total_responses]]
            chart_labels = ['Total Forms', 'Active Forms', 'Total Responses']

            # Dynamic scaling for better visuals
            max_value = max(total_forms, active_forms, total_responses, 1)
            value_max = max_value + 1

            # Drawing area
            drawing = Drawing(500, 300)

            # Chart setup
            bc = VerticalBarChart()
            bc.x = 70
            bc.y = 70
            bc.width = 350
            bc.height = 150
            bc.data = chart_data
            bc.barWidth = 35

            # X-axis labels
            bc.categoryAxis.categoryNames = chart_labels
            bc.categoryAxis.labels.boxAnchor = 'n'
            bc.categoryAxis.labels.fontName = 'Helvetica-Bold'
            bc.categoryAxis.labels.fontSize = 10

            # Y-axis labels
            bc.valueAxis.valueMin = 0
            bc.valueAxis.valueMax = value_max
            bc.valueAxis.valueStep = max(1, round(value_max / 5))
            bc.valueAxis.labels.fontName = 'Helvetica'
            bc.valueAxis.labels.fontSize = 9

            # Bar colors
            bc.bars[0].fillColor = colors.HexColor('#5B9BD5')

            # Title
            title = String(160, 250, "Summary Statistics", fontSize=16, fontName="Helvetica-Bold")
            drawing.add(title)

            # X-axis label
            x_label = String(220, 15, "Metrics", fontSize=12, fontName="Helvetica-Bold")
            drawing.add(x_label)

            # Y-axis label (rotated)
            y_text = String(220, 10, "Count", fontSize=12, fontName="Helvetica-Bold")
            # y_text.textAnchor = 'middle'
            y_label = Group(y_text)
            y_label.rotate(90)
            y_label.translate(160, 250)
            drawing.add(y_label)

            # Legend
            legend = Legend()
            legend.x = 440
            legend.y = 160
            legend.dx = 10
            legend.dy = 10
            legend.fontName = 'Helvetica'
            legend.fontSize = 10
            legend.boxAnchor = 'w'
            legend.colorNamePairs = [(colors.HexColor('#5B9BD5'), 'Series1')]
            drawing.add(legend)

            drawing.add(bc)
            story.append(Spacer(1, 30))
            story.append(drawing)
            



            # Question analytics section (summary)
            story.append(Paragraph("Question Analytics Summary", overview_style))

            total_questions = 0
            total_responses = 0
            for form in forms:
                total_questions += form.questions.count()
                total_responses += FeedbackResponse.objects.filter(form=form).count()

            summary_data = [
                ['Total Questions', str(total_questions)],
                ['Total Responses', str(total_responses)],
                ['Average Responses per Form', f"{total_responses / forms.count():.1f}" if forms.count() > 0 else "0"]
            ]

            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            story.append(summary_table)

            # Build PDF
            doc.build(story)
            return response

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Unable to export analytics: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DashboardView(APIView):
    """Dashboard view for admin overview"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user

        # If superuser — allow filtering by ?admin_id=
        if user.is_superuser:
            admin_name = request.query_params.get('admin_name')
            if admin_name:
                try:
                    admin_user = User.objects.get(username=admin_name)
                except User.DoesNotExist:
                    return Response({"detail": "Admin not found."}, status=404)
                queryset_user = admin_user
            else:
                queryset_user = None  # Means all admins

            if queryset_user:
                forms = FeedbackForm.objects.filter(created_by=queryset_user)
            else:
                forms = FeedbackForm.objects.all()

            total_responses = FeedbackResponse.objects.filter(
                form__in=forms
            ).count()

            day_ago = timezone.now() - timedelta(hours=24)
            recent_responses = FeedbackResponse.objects.filter(
                form__in=forms,
                submitted_at__gte=day_ago
            ).count()

            analytics = FormAnalytics.objects.filter(form__in=forms)
            avg_completion_rate = analytics.aggregate(
                avg_rate=Avg('completion_rate')
            )['avg_rate'] or 0.0

            recent_response_data = FeedbackResponse.objects.filter(
                form__in=forms
            ).select_related('form').order_by('-submitted_at')[:10]

        # 1 If normal admin — only their forms
        else:
            forms = FeedbackForm.objects.filter(created_by=user)
            total_responses = FeedbackResponse.objects.filter(
                form__created_by=user
            ).count()

            day_ago = timezone.now() - timedelta(hours=24)
            recent_responses = FeedbackResponse.objects.filter(
                form__created_by=user,
                submitted_at__gte=day_ago
            ).count()

            analytics = FormAnalytics.objects.filter(form__created_by=user)
            avg_completion_rate = analytics.aggregate(
                avg_rate=Avg('completion_rate')
            )['avg_rate'] or 0.0

            recent_response_data = FeedbackResponse.objects.filter(
                form__created_by=user
            ).select_related('form').order_by('-submitted_at')[:10]

        # 📝 Recent response list
        recent_responses_list = [
            {
                'id': str(response.id),
                'form_title': response.form.title,
                'submitted_at': response.submitted_at,
                'form_id': str(response.form.id)
            }
            for response in recent_response_data
        ]

        # 📊 Summary
        summary_data = {
            'total_forms': forms.count(),
            'active_forms': forms.filter(is_active=True).count(),
            'total_responses': total_responses,
            'recent_responses': recent_responses,
            'average_completion_rate': round(avg_completion_rate, 2),
            'recent_responses_list': recent_responses_list
        }

        serializer = FormSummarySerializer(summary_data)
        return Response(serializer.data)


class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for managing notifications"""
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user)
    
    @action(detail=True, methods=['post'])
    def mark_as_read(self, request, pk=None):
        """Mark a notification as read"""
        notification = self.get_object()
        notification.is_read = True
        notification.save()
        return Response({'status': 'marked as read'})
    
    @action(detail=False, methods=['post'])
    def mark_all_as_read(self, request):
        """Mark all notifications as read"""
        self.get_queryset().update(is_read=True)
        return Response({'status': 'all marked as read'})
    
    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        """Get count of unread notifications"""
        count = self.get_queryset().filter(is_read=False).count()
        return Response({'unread_count': count})


class CustomAuthToken(ObtainAuthToken):
    """Custom authentication view that returns user data along with token"""
    permission_classes = [AllowAny]
    
    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']

        # 🛡️ Add approval check here
        if not user.is_approved and not user.is_superuser:
            return Response(
                {"detail": "Your account is awaiting approval by the admin."},
                status=status.HTTP_403_FORBIDDEN
            )

        token, created = Token.objects.get_or_create(user=user)
        
        return Response({
            'token': token.key,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'is_staff': user.is_staff,
                'is_superuser': user.is_superuser,
            }
        })


class LogoutView(APIView):
    """Logout view to invalidate token"""
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        try:
            # Delete the token
            request.user.auth_token.delete()
        except:
            pass
        
        # Logout from Django session
        django_logout(request)
        
        return Response({'message': 'Successfully logged out'})


class CurrentUserView(APIView):
    """Get current user information"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        user = request.user
        return Response({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
        })
    
# class FormAnalyticsView(APIView):
#     permission_classes = [permissions.IsAuthenticated]

#     def get(self, request, form_id):
#         try:
#             form = FeedbackForm.objects.get(id=form_id)
#             analytics, created = FormAnalytics.objects.get_or_create(form=form)
#             analytics.update_analytics()  # Ensure data is up-to-date
#             serializer = FormAnalyticsSerializer(analytics)
#             return Response(serializer.data)
#         except FeedbackForm.DoesNotExist:
#             return Response({"detail": "Form not found."}, status=404)
        
